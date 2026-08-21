#!/usr/bin/env python3
"""Generate an Excel (.xlsx) spreadsheet from text input."""
import sys, os, csv, io, re

def detect_and_parse(text):
    lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
    if not lines: return []
    try:
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if len(rows) >= 2 and len(rows[0]) >= 2: return rows
    except: pass
    if all('\t' in line for line in lines):
        rows = [[c.strip() for c in line.split('\t')] for line in lines]
        if len(rows) >= 2 and len(rows[0]) >= 2: return rows
    if all('|' in line for line in lines):
        rows = [[c.strip() for c in line.split('|') if c.strip()] for line in lines]
        if len(rows) >= 2 and len(rows[0]) >= 2: return rows
    if all(',' in line for line in lines):
        rows = [[c.strip() for c in line.split(',')] for line in lines]
        if len(rows) >= 2 and len(rows[0]) >= 2: return rows
    if len(lines) >= 3:
        rows = [re.split(r'\s{2,}', l.strip()) for l in lines]
        if len(rows) >= 2 and len(rows[0]) >= 2 and len(set(len(r) for r in rows)) == 1: return rows
    return [[line] for line in lines]

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    if len(sys.argv) < 3:
        print("Usage: python3 generate_xlsx.py <input_text_file> <output_path> [title]", file=sys.stderr)
        sys.exit(1)

    text_file, output_path = sys.argv[1], sys.argv[2]
    title = sys.argv[3] if len(sys.argv) > 3 else "Spreadsheet"

    with open(text_file, 'r', encoding='utf-8') as f:
        text = f.read()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Data"

    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='16213E', end_color='16213E', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Calibri', size=11)
    cell_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

    rows = detect_and_parse(text)
    if rows:
        for ci, v in enumerate(rows[0], 1):
            cell = ws.cell(row=1, column=ci, value=v)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = thin_border
        for ri, row_data in enumerate(rows[1:], 2):
            for ci, v in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
                if ri % 2 == 0:
                    cell.fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
        for ci in range(1, len(rows[0]) + 1):
            cl = get_column_letter(ci)
            mw = max(len(str(rows[r][ci-1])) for r in range(min(len(rows), 50)))
            ws.column_dimensions[cl].width = max(mw + 2, 12)
        ws.freeze_panes = 'A2'
    else:
        ws.cell(row=1, column=1, value=text)

    wb.save(output_path)
    print(output_path)
except ImportError as e:
    print(f"Error importing dependencies: {e}", file=sys.stderr); sys.exit(1)
except Exception as e:
    print(f"Error generating Excel: {e}", file=sys.stderr); sys.exit(1)
